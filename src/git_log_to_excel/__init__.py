import re
from re import Pattern
import subprocess
import pandas as pd
from pathlib import Path

DATETIME_PATTERN: Pattern = re.compile(r"^(\d{4}-\d{2}-\d{2}) \d{2}\:\d{2}\:\d{2} \+\d{4}")
CAT_PATTERN = re.compile(r"\d{4}\: (fix|feat|refactor|perf)")
CATEGORY: dict[str, str] = {
    "feat": "新增需求",
    "fix": "bug修复",
    "refactor": "功能改进",
    "perf": "功能改进",
}
CAT_ORDER = {
    "新增需求": 0,
    "bug修复": 1,
    "功能改进": 2,
}
COLUMN_ALIAS = {
    "cat": "需求分类",
    "content": "需求说明",
    "date": "实际完成时间",
    "module": "功能模块",
    "date_delta": "工作量(天)",
}


def match_date(raw: str) -> str | None:
    result = DATETIME_PATTERN.match(raw)
    return result.groups()[0] if result is not None else pd.NA


def match_category(raw: str) -> str | None:
    result = CAT_PATTERN.findall(raw)
    if len(result) == 0:
        return pd.NA
    text_match = result[0]
    rename = CATEGORY.get(text_match)
    return rename if rename is not None else pd.NA


def get_content(raw: str) -> str | None:
    raw = raw.replace("\n", "").split(":")[-1]
    raw = raw.strip()
    return raw if raw != "" else pd.NA


from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment


def write_to_excel(
        output_path: str | Path,
        rows: list[list[str]],
        columns: list[str],
) -> None:
    thin_border = Border(left=Side(style="thin"),
                         right=Side(style="thin"),
                         top=Side(style="thin"),
                         bottom=Side(style="thin"))
    left_border = Border(left=Side(style="thin"))

    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    try:
        cat_index = columns.index("cat")
        module_index = columns.index("module")
        renamed_columns = [COLUMN_ALIAS.get(x) for x in columns]
        ws.append(renamed_columns)
        [ws.append(row) for row in rows]
        for i in range(1, len(columns) + 1):
            for j in range(1, len(rows) + 2):
                cell_ref = ws.cell(row=j, column=i)
                cell_ref.border = thin_border
                if i == 2:
                    cell_ref.alignment = Alignment(vertical="center", horizontal="left")
                else:
                    cell_ref.alignment = Alignment(vertical="center", horizontal="center")

        mod_col = module_index + 1
        cat_col = cat_index + 1
        m_start = 2
        for r in range(2, len(rows) + 3):
            current_mod = ws.cell(row=r, column=mod_col).value
            prev_mod = ws.cell(row=r - 1, column=mod_col).value
            if current_mod != prev_mod or r > len(rows) + 1:
                m_end = r - 1
                if m_end > m_start:
                    ws.merge_cells(start_row=m_start, start_column=mod_col, end_row=m_end, end_column=mod_col)
                c_start = m_start
                for cr in range(m_start, m_end + 2):
                    current_cat = ws.cell(row=cr, column=cat_col).value
                    prev_cat = ws.cell(row=cr - 1, column=cat_col).value
                    if (current_cat != prev_cat and cr > m_start) or cr > m_end:
                        c_end = cr - 1
                        if c_end > c_start:
                            ws.merge_cells(start_row=c_start, start_column=cat_col, end_row=c_end, end_column=cat_col)
                        c_start = cr
                m_start = r
        wb.save(output_path)
        print("写入成功!")
    except Exception as e:
        print("写到XLSX时异常: " + e)
    finally:
        wb.close()


def log_collector(
        module_path: dict[str, Path],
        start_date: str,
        end_date: str,
        output_path: Path | str,
) -> dict | None:
    cmd = [
        "log",
        f'--since={start_date}',
        f'--until={end_date}',
        '--pretty=format:"%ad: %s"',
        "--date=iso"
    ]
    rows = []
    columns = None
    for module, path in module_path.items():
        if isinstance(path, str):
            path = Path(path)
        if not isinstance(path, Path):
            print(f"路径数据类型必须是str, value: {path}, type: {type(path)}")
        if not path.exists():
            print(f"路径 {str(path.absolute())} 不存在!")
            return None
        try:
            this_cmd = ["git", "-C", f'{str(path.absolute())}'] + cmd
            result = subprocess.run(this_cmd,
                                    capture_output=True,
                                    text=True,
                                    encoding='utf-8',  # 指定 UTF-8 编码
                                    check=True)
            git_log = result.stdout
        except subprocess.CalledProcessError as e:
            print(f"获取日志时发生错误: {e.stderr}")
            break
        if len(git_log) == 0:
            continue
        module_df = pd.DataFrame(
            data={
                "ori": git_log.replace('"', "").split("\n")
            }
        )
        module_df["date"] = module_df["ori"].apply(match_date)
        module_df["cat"] = module_df["ori"].apply(match_category)
        module_df = module_df.loc[module_df["cat"].notna() & module_df["date"].notna()]
        module_df["content"] = module_df["ori"].apply(get_content)
        module_df["date_temp"] = pd.to_datetime(module_df["date"])
        module_df = module_df.sort_values(by=['content', 'date_temp'])
        module_df["date_delta"] = module_df.groupby('content')['date_temp'].transform(lambda x: (x - x.min()).dt.days)
        module_df["date_delta"] = module_df['date_delta'].apply(lambda x: 1 if x == 0 else x)
        module_df = module_df.drop_duplicates(subset=["cat", "content"], keep="last")
        module_df = module_df.loc[module_df["content"].notna()]
        module_df["cat"] = pd.Categorical(module_df['cat'], categories=CAT_ORDER.keys(), ordered=True)
        module_df = module_df.sort_values(by=["cat", "date_temp"], ascending=[True, False])
        module_df["module"] = module
        module_df = module_df[["module", "content", "cat", "date", "date_delta"]]
        module_df = module_df.reset_index(drop=True)
        module_recs = module_df.to_dict(orient="split")
        if columns is None:
            columns = module_recs["columns"]
        rows.extend(module_recs["data"])
    if rows and columns is not None:
        write_to_excel(output_path, rows, columns)
    else:
        print("无数据可写入!")
